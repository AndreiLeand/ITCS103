import tkinter as tk
import openpyxl as op
from tkinter import ttk, messagebox

Lab_schedule = {
    "Monday": {},
    "Tuesday": {},
    "Wednesday": {},
    "Thursday": {},
    "Friday": {},
}
def show_welcome_window():
    # Hide the main window first
    window.withdraw()
    
    # Create welcome window
    welcome_win = tk.Toplevel(window)
    welcome_win.title("Welcome")
    welcome_win.geometry("600x400")
    welcome_win.configure(bg="#2c3e50")
    welcome_win.resizable(False, False)
    
    # Center the window
    welcome_win.update_idletasks()
    width = 600
    height = 400
    x = (welcome_win.winfo_screenwidth() // 2) - (width // 2)
    y = (welcome_win.winfo_screenheight() // 2) - (height // 2)
    welcome_win.geometry(f'{width}x{height}+{x}+{y}')
    
    # Title
    title_label = tk.Label(welcome_win, 
                          text="COMPUTER LABORATORY\nRESERVATION SYSTEM", 
                          font=("Times New Roman", 24, "bold"), 
                          fg="white", 
                          bg="#2c3e50")
    title_label.pack(pady=30)
    
    # Description
    desc_text = """
    Welcome to the DLL Computer Laboratory Reservation System!
    
    This system allows authorized student representatives to:
    
    • Reserve computer laboratories (CL1-CL5)
    • View existing schedules
    • Update and delete reservations
    
    Please ensure you are an authorized representative
    before making a reservation.
    """
    
    desc_label = tk.Label(welcome_win, 
                         text=desc_text, 
                         font=("Arial", 11), 
                         fg="#ecf0f1", 
                         bg="#2c3e50",
                         justify="center")
    desc_label.pack(pady=10)
    
    # Continue Button
    def proceed_to_main():
        welcome_win.destroy()  # Close welcome window
        window.deiconify()     # Show main window
        window.state('zoomed') # Make main window fullscreen
    
    btn_continue = tk.Button(welcome_win, 
                            text="PROCEED TO SYSTEM", 
                            font=("Arial", 12, "bold"), 
                            bg="#27ae60", 
                            fg="white", 
                            width=20, 
                            command=proceed_to_main)
    btn_continue.pack(pady=20)


def checking_schedule():
    global Lab_schedule
    
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    labs = ["CL1", "CL2", "CL3", "CL4", "CL5"]
    times = ["07:00 - 08:30am", "08:30 - 10:00am", "10:00 - 12:00NN", 
             "01:00 - 02:00pm", "02:00 - 03:30pm", "03:30 - 05:00pm", "05:00 - 07:30pm"]
    
    for day in days:
        Lab_schedule[day] = {}
        for lab in labs:
            Lab_schedule[day][lab] = {}
            for time_slot in times:
                Lab_schedule[day][lab][time_slot] = []
    
    try:
        workbook = op.load_workbook("Menemedez_Database.xlsx")
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            lab = row[6]
            day = row[7]
            time = row[8]
            course = row[3]
            year = str(row[4])
            section = row[5]
            
            full_section = f"{course}-{year}-{section}"
            
            if day in Lab_schedule:
                if lab in Lab_schedule[day]:
                    if time in Lab_schedule[day][lab]:
                        Lab_schedule[day][lab][time].append(full_section)
    except:
        pass  # Ignore errors

# def view_all_schedules():
#     # Create a new window 
#     schedule_window = tk.Toplevel(window)
#     schedule_window.title("Laboratory Schedules")
#     schedule_window.geometry("1000x700")
#     schedule_window.configure(bg="white")
    
#     # Title
#     title = tk.Label(schedule_window, 
#                      text="COMPUTER LABORATORY SCHEDULES", 
#                      font=("Times New Roman", 16, "bold"),
#                      bg="white")
#     title.pack(pady=10)
    
#     # Create a frame for all labs
#     labs_frame = tk.Frame(schedule_window, bg="white")
#     labs_frame.pack(fill="both", expand=True, padx=20, pady=10)
    
#     # Define labs
#     labs = ["CL1", "CL2", "CL3", "CL4", "CL5"]
#     days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
#     times = ["07:00 - 08:30am", "08:30 - 10:00am", "10:00 - 12:00NN", 
#              "01:00 - 02:00pm", "02:00 - 03:30pm", "03:30 - 05:00pm", "05:00 - 07:30pm"]
    
#     # Create a notebook (tabs) for each laboratory
#     from tkinter import ttk
#     notebook = ttk.Notebook(labs_frame)
#     notebook.pack(fill="both", expand=True)
    
#     # Create a tab for each lab
#     for lab in labs:
#         lab_frame = tk.Frame(notebook, bg="white")
#         notebook.add(lab_frame, text=lab)
        
#         # Create Treeview for this lab
#         columns = ("Day", "Time", "Course", "Section")
#         tree = ttk.Treeview(lab_frame, columns=columns, show="headings", height=15)
        
#         # Set headings
#         tree.heading("Day", text="Day")
#         tree.heading("Time", text="Time")
#         tree.heading("Course", text="Course")
#         tree.heading("Section", text="Section")
        
#         # Set column widths
#         tree.column("Day", width=100)
#         tree.column("Time", width=150)
#         tree.column("Course", width=100)
#         tree.column("Section", width=100)
        
#         # Add scrollbar
#         scrollbar = ttk.Scrollbar(lab_frame, orient="vertical", command=tree.yview)
#         tree.configure(yscrollcommand=scrollbar.set)
        
#         tree.pack(side="left", fill="both", expand=True)
#         scrollbar.pack(side="right", fill="y")
        
#         # Populate with data from Lab_schedule dictionary
#         for day in days:
#             for time in times:
#                 # Check if this slot has occupants
#                 if lab in Lab_schedule.get(day, {}):
#                     occupants = Lab_schedule[day][lab].get(time, [])
#                     if occupants:
#                         for occupier in occupants:
#                             # Parse the occupier string (e.g., "BSIT-1-A")
#                             parts = occupier.split("-")
#                             if len(parts) >= 3:
#                                 course = parts[0]
#                                 year = parts[1]
#                                 section = parts[2]
#                                 tree.insert("", "end", values=(day, time, course, f"{year}-{section}"))
        
#         # If no schedules, show a message
#         if tree.get_children() == ():
#             no_data_label = tk.Label(lab_frame, 
#                                      text="No reservations for this laboratory",
#                                      font=("Arial", 10, "italic"),
#                                      bg="white", fg="gray")
#             no_data_label.pack(pady=20)        



def display_excel():
    workbook = op.load_workbook("Menemedez_Database.xlsx")
    sheet = workbook.active

    for content in tree.get_children():
        tree.delete(content)

    for row in sheet.iter_rows(min_row=2,values_only=True):
        tree.insert("",tk.END,values=row)

def validation():
    firstname = fname_entry.get()
    lastname = lname_entry.get()
    course = course_entry.get()
    year = year_entry.get()
    section = section_entry.get()


    if not firstname or not lastname or not course or not year or not section:
        messagebox.showerror("Error","All fields is required")
        return False
    if not year.isdigit():
        messagebox.showerror("Error","Year must be a number")
        return False
    return True



def reserve():
    # Get all inputs
    firstname = fname_entry.get()
    lastname = lname_entry.get()
    course = course_entry.get()
    year = year_entry.get()
    section = section_entry.get()
    lab = lab_list.get()
    day = day_list.get()
    time = time_list.get()
    
    if not firstname or not lastname or not course or not year or not section:
        messagebox.showerror("Error", "Please fill all fields!")
        return
    
    if not year.isdigit():
        messagebox.showerror("Error", "Year must be a number!")
        return
    
    try:
        if Lab_schedule[day][lab][time]:
            occupants = ", ".join(Lab_schedule[day][lab][time])
            messagebox.showerror("Reservation Failed", 
                                f"Cannot Reserve!\n\n{lab} on {day} at {time} is already occupied by:\n{occupants}")
            return
    except KeyError:
        messagebox.showerror("Error", "Invalid Schedule Selection.")
        return
    
    workbook = op.load_workbook("Menemedez_Database.xlsx")
    sheet = workbook.active
    
    new_id = sheet.max_row
    
    sheet.append([new_id, lastname, firstname, course, year, section, lab, day, time])
    
    workbook.save("Menemedez_Database.xlsx")
    
    full_section = f"{course}-{year}-{section}"
    Lab_schedule[day][lab][time].append(full_section)
    
    messagebox.showinfo("Success", "Reservation successful!")
    
    display_excel()
    
    clear_form()


def clear_form():
    fname_entry.delete(0, tk.END)
    lname_entry.delete(0, tk.END)
    course_entry.delete(0, tk.END)
    year_entry.delete(0, tk.END)
    section_entry.delete(0, tk.END)
    
    lab_list.current(0)
    day_list.current(0)
    time_list.current(0)


def select_for_update(event):
    selected_item = tree.focus()
    if not selected_item:
        return
    
    values = tree.item(selected_item)['values']
    
    clear_form()
    
    
    fname_entry.insert(0, values[2])  # First Name
    lname_entry.insert(0, values[1])  # Last Name
    course_entry.insert(0, values[3]) # Course
    year_entry.insert(0, values[4])   # Year
    section_entry.insert(0, values[5])# Section
    
    lab_list.set(values[6])
    day_list.set(values[7])
    time_list.set(values[8])

    
    global editing_item_id
    editing_item_id = selected_item    


def update_record():
    selected_item = tree.focus()
    
    if not selected_item:
        messagebox.showwarning("Warning", "Please select a record from the table to update.")
        return

    firstname = fname_entry.get().strip()
    lastname = lname_entry.get().strip()
    course = course_entry.get().strip()
    year = year_entry.get().strip()
    section = section_entry.get().strip()
    lab = lab_list.get()
    day = day_list.get()
    time = time_list.get()

    if not all([firstname, lastname, course, year, section]):
        messagebox.showerror("Error", "All fields must be filled for update.")
        return

    try:
        workbook = op.load_workbook("Menemedez_Database.xlsx")
        sheet = workbook.active
        
        target_id = tree.item(selected_item)['values'][0]
        
        found = False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == target_id: # Column A is ID
                row[1].value = lastname
                row[2].value = firstname
                row[3].value = course
                row[4].value = int(year)
                row[5].value = section
                row[6].value = lab
                row[7].value = day
                row[8].value = time
                found = True
                break
        
        if not found:
            messagebox.showerror("Error", "Record ID not found in database.")
            return
        
        try:
            if Lab_schedule[day][lab][time]:
                occupants = ", ".join(Lab_schedule[day][lab][time])
                messagebox.showerror("Reservation Failed", 
                                    f"Cannot Reserve!\n\n{lab} on {day} at {time} is already occupied by:\n{occupants}")
                return
        except KeyError:
            messagebox.showerror("Error", "Invalid Schedule Selection.")
            return
        workbook.save("Menemedez_Database.xlsx")

        checking_schedule() 

        display_excel()
        clear_form()
        messagebox.showinfo("Success", "Record Updated Successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to update: {str(e)}")


def delete_record():
    selected_item = tree.focus()
    
    if not selected_item:
        messagebox.showwarning("Warning", "Please select a record to delete.")
        return

    confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this reservation?")
    if not confirm:
        return

    try:
        target_id = tree.item(selected_item)['values'][0]

        workbook = op.load_workbook("Menemedez_Database.xlsx")
        sheet = workbook.active
        
        for row in reversed(list(sheet.iter_rows(min_row=2))):
            if row[0].value == target_id:
                sheet.delete_rows(row[0].row)
                break
                
        workbook.save("Menemedez_Database.xlsx")

        checking_schedule() 
        display_excel()     
        clear_form()        
        
        messagebox.showinfo("Success", "Record Deleted Successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to delete: {str(e)}")

window = tk.Tk()
window.title("COMPUTER LABORATORY RESERVATION SYSTEM")
window.geometry("900x600")
window.configure(bg="lightblue")

# Configure the main window grid to allow expansion
window.grid_rowconfigure(4, weight=1) # Allows the treeview to expand vertically
window.grid_columnconfigure(0, weight=1)

# --- TITLE ---
title = tk.Label(window, text="COMPUTER LABORATORY RESERVATION SYSTEM", font=("Times New Roman", 16, "bold"), bg="lightblue")
title.grid(row=0, column=0, pady=10)

# --- INPUT FRAME (Personal Info) ---
frame = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

# ROW 0: First Name & Last Name
fname = tk.Label(frame, text="First Name", font=("Poppins", 10, "italic"), bg="lightblue")
fname.grid(row=0, column=0, padx=5, pady=5, sticky="e")
fname_entry = tk.Entry(frame, font=("Poppins", 12), width=20)
fname_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

lname = tk.Label(frame, text="Last Name", font=("Poppins", 10, "italic"), bg="lightblue")
lname.grid(row=0, column=2, padx=5, pady=5, sticky="e")
lname_entry = tk.Entry(frame, font=("Poppins", 12), width=20)
lname_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")

# ROW 1: Course, Year, Section
course = tk.Label(frame, text="Course", font=("Poppins", 10, "italic"), bg="lightblue")
course.grid(row=1, column=0, padx=5, pady=5, sticky="e")
course_entry = tk.Entry(frame, font=("Poppins", 12), width=20)
course_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

year = tk.Label(frame, text="Year", font=("Poppins", 10, "italic"), bg="lightblue")
year.grid(row=1, column=2, padx=5, pady=5, sticky="e")
year_entry = tk.Entry(frame, font=("Poppins", 12), width=15)
year_entry.grid(row=1, column=3, padx=5, pady=5, sticky="w")

section = tk.Label(frame, text="Section", font=("Poppins", 10, "italic"), bg="lightblue")
section.grid(row=1, column=4, padx=5, pady=5, sticky="e")
section_entry = tk.Entry(frame, font=("Poppins", 12), width=15)
section_entry.grid(row=1, column=5, padx=5, pady=5, sticky="w")

# --- RESERVATION DETAILS (Lab & Day using Comboboxes) ---
res_frame = tk.Frame(window, bg="lightblue")
res_frame.grid(row=2, column=0, pady=15)

# 1. SELECT LABORATORY
select_lab = tk.Label(res_frame, text="Select Laboratory:", font=("Poppins", 10, "bold"), bg="lightblue")
select_lab.grid(row=0, column=0, padx=10, sticky="e")

lab_list = ttk.Combobox(res_frame, values=["","CL1", "CL2", "CL3", "CL4", "CL5"], state="readonly", font=("Poppins", 12), width=15)
lab_list.grid(row=0, column=1, padx=10)
lab_list.current(0) # Default to CL1

# 2. SELECT DAY
select_day = tk.Label(res_frame, text="Select Schedule:", font=("Poppins", 10, "bold"), bg="lightblue")
select_day.grid(row=0, column=2, padx=10, sticky="e")

day_list = ttk.Combobox(res_frame, values=["","Monday", "Tuesday", "Wednesday", "Thursday", "Friday"], state="readonly", font=("Poppins", 12), width=15)
day_list.grid(row=0, column=3, padx=10)
day_list.current(0) # Default to MWF

# 3. SELECT TIME

select_time = tk.Label(res_frame,text="Select Time:",font=("Poppins",10,"bold"),bg="lightblue")
select_time.grid(row=0,column=4,padx=10,sticky="e")
time_list = ttk.Combobox(res_frame,values=["",
                                        "07:00 - 08:30am",
                                           "08:30 - 10:00am",
                                           "10:00 - 12:00NN",
                                           "01:00 - 02:00pm",
                                           "02:00 - 03:30pm",
                                           "03:30 - 05:00pm",
                                           "05:00 - 07:30pm"],state="readonly",font=("Poppins",12),width=15)
time_list.grid(row=0,column=5,padx=10)
time_list.current(0)



# --- ACTION BUTTONS ---
btn_frame = tk.Frame(window, bg="lightblue")
btn_frame.grid(row=3, column=0, pady=5)
# Add this button (put it before or after your other buttons)

# btn_view_schedule = tk.Button(btn_frame, 
#                                text="VIEW ALL SCHEDULES", 
#                                width=20, 
#                                height=2, 
#                                font=("Arial", 10, "bold"), 
#                                bg="#009688", 
#                                fg="white",
#                                command=view_all_schedules)
# btn_view_schedule.grid(row=0, column=4, padx=10)

btn_reserve = tk.Button(btn_frame, text="RESERVE SLOT", width=15, height=2, font=("Arial", 10, "bold"), bg="#4CAF50", fg="white",command=reserve)
btn_reserve.grid(row=0, column=0, padx=10)

btn_update = tk.Button(btn_frame, text="UPDATE", width=15, height=2, font=("Arial", 10, "bold"), bg="#330be7", fg="white",command=update_record)
btn_update.grid(row=0, column=1, padx=10)

btn_delete = tk.Button(btn_frame, text="DELETE", width=15, height=2, font=("Arial", 10, "bold"), bg="#f44336", fg="white",command=delete_record)
btn_delete.grid(row=0, column=2, padx=10)

btn_clear = tk.Button(btn_frame, text="CLEAR FORM", width=15, height=2, font=("Arial", 10, "bold"), bg="#9E9E9E", fg="white",command=clear_form)
btn_clear.grid(row=0, column=3, padx=10)



# Treeview Setup
columns = ("ID", "Last Name", "First Name", "Course", "Year", "Section", "Lab", "Day","Time")
tree = ttk.Treeview(window, columns=columns,show="headings")
for col in columns:
    tree.heading(col,text=col)
    tree.column(col,width=100)
tree.grid(row=7,column=0,columnspan=4)
tree.bind("<<TreeviewSelect>>",select_for_update)

checking_schedule()
display_excel()
show_welcome_window()

window.mainloop()

