import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

#FUNCTION FOR BACKEND


def display_excel():
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    #Clear Treeview
    for content in table.get_children():
        table.delete(content)

    #Insert Excel Data
    for row in sheet.iter_rows(min_row=2,values_only=True):
        table.insert("",tk.END,values=row)

def select_record():
    selected = table.focus()
    values = table.item(selected, "values")
    
    if not values:
        return
    
    # Clear current entries
    fname_entry.delete(0, tk.END)
    mname_entry.delete(0, tk.END)
    lname_entry.delete(0, tk.END)
    birth_entry.delete(0, tk.END)
    
    # Insert selected values into entries
    # Note: Order in Excel is [ID, Last, First, Middle, BirthYear, Age]
    # So we skip ID and Age for input fields
    lname_entry.insert(0, values[1])      # Last Name
    fname_entry.insert(0, values[2])      # First Name
    mname_entry.insert(0, values[3])      # Middle Name
    birth_entry.insert(0, str(values[4])) # Birth Year
    
    

def update_data():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to update.")
        return
    
    values = table.item(selected, "values")
    record_id = int(values[0])  # ID is first column
    
    if not validation():
        return
    
    first = fname_entry.get().strip()
    middle = mname_entry.get().strip()
    last = lname_entry.get().strip()
    try:
        birthyear = int(birth_entry.get())
    except ValueError:
        messagebox.showerror("Error", "Birth year must be a number.")
        return
    
    age = 2026 - birthyear
    
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active
    
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id:
            row[1].value = last
            row[2].value = first
            row[3].value = middle
            row[4].value = birthyear
            row[5].value = age
            found = True
            break
    
    if found:
        workbook.save("excelDB.xlsx")
        messagebox.showinfo("Success", "Record updated successfully!")
        display_excel()
    else:
        messagebox.showerror("Error", "Record not found.")



def validation():
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birthyear = birth_entry.get()

    if not first or not middle or not last or not birthyear:
        messagebox.showerror("Error","All fields is required")
        return False
    if not birthyear.isdigit():
        messagebox.showerror("Error","Birth year must be a number")
        return False
    return True

def append_excel():
    if not validation():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birthyear = int(birth_entry.get())
    age = 2026 - birthyear
    
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id,last,first,middle,birthyear,age])
    workbook.save("excelDB.xlsx")

    messagebox.showinfo("Success","Record added successfully!")
    display_excel()

def delete_data():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to delete.")
        return
    
    values = table.item(selected, "values")
    record_id = int(values[0])
    
    confirm = messagebox.askyesno("Confirm", f"Are you sure you want to delete record ID {record_id}?")
    if not confirm:
        return
    
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active
    
    deleted = False
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == record_id:
            sheet.delete_rows(i, 1)
            deleted = True
            break
    
    if deleted:
        workbook.save("excelDB.xlsx")
        messagebox.showinfo("Success", "Record deleted successfully!")
        display_excel()
    else:
        messagebox.showerror("Error", "Record not found.")



window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

update_btn = tk.Button(window, text="Update",command=update_data)
update_btn.grid(row=6, column=2)

submit_button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=append_excel)
submit_button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white",command=delete_data)
delete_btn.grid(row=6, column=3)

columns = ("ID", "Last", "First", "Middle", "BirthYear", "Age")
table = ttk.Treeview(window, columns=columns, show="headings")
for col in columns:
    table.heading(col, text=col)
table.grid(row=7, column=0, columnspan=4)
table.bind("<<TreeviewSelect>>",select_record)

display_excel() 


window.mainloop()